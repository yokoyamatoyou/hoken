import base64
import datetime
import json
import logging
import os
import queue
import re
import shutil
import threading
import functools

import customtkinter as ctk
import tkinter
from tkinter import filedialog, messagebox
from PIL import Image
import docx
import PyPDF2
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI

from src.agent import ReActAgent, CoTAgent, ToTAgent, PresentationAgent
from src.main import create_evaluator, read_tot_env
from src.constants import TOT_LEVELS
from src.memory import ConversationMemory
from src.tools import (
    get_web_scraper,
    get_sqlite_tool,
    get_graphviz_tool,
    get_mermaid_tool,
)
from src.tools.graphviz_tool import create_graphviz_diagram
from src.tools.mermaid_tool import create_mermaid_diagram, sanitize_mermaid_code


def get_font_family(preferred: str = "Meiryo") -> str:
    """Return an available font family.

    The ``PREFERRED_FONT`` environment variable can specify a single font or a
    comma-separated list of candidates. The first available font is used before
    falling back to the ``preferred`` parameter and finally ``Helvetica``.
    """
    env_font = os.getenv("PREFERRED_FONT")
    candidates = []
    if env_font:
        candidates.extend([f.strip() for f in env_font.split(",") if f.strip()])
    candidates.append(preferred)
    try:
        root = tkinter.Tk()
        root.withdraw()
        families = set(root.tk.call("font", "families"))
        root.destroy()
        for font in candidates:
            if font in families:
                return font
    except tkinter.TclError:
        pass
    return "Helvetica"



# Mapping of tool names to implementation functions
TOOL_FUNCS = {
    "create_graphviz_diagram": create_graphviz_diagram,
    "create_mermaid_diagram": create_mermaid_diagram,
}

# Load environment variables from .env if present
load_dotenv()

# Default directory for saving conversations
CONV_DIR = os.getenv("CONVERSATION_DIR", "conversations")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# CustomTkinterの設定
ctk.set_appearance_mode("light")
GOOGLE_THEME = os.path.join(os.path.dirname(__file__), "resources", "google.json")
ctk.set_default_color_theme(GOOGLE_THEME)

ICON_PATH = os.path.join(os.path.dirname(__file__), "resources", "app_icon.xbm")

FONT_FAMILY = get_font_family()

# Custom frame that pulls its colors from the "DiagramFrame" theme entry.
if hasattr(ctk, "CTkFrame"):
    class DiagramFrame(ctk.CTkFrame):
        def __init__(self, *args, **kwargs):
            if kwargs.get("fg_color") is None and hasattr(ctk, "ThemeManager"):
                theme = ctk.ThemeManager.theme.get("DiagramFrame", {})
                color = theme.get("fg_color")
                if color is not None:
                    kwargs["fg_color"] = color
            super().__init__(*args, **kwargs)
else:
    class DiagramFrame(object):  # type: ignore
        def __init__(self, *args, **kwargs):
            pass

# Preset search parameters for the Tree-of-Thoughts agent are defined in
# :mod:`src.constants` as ``TOT_LEVELS``.

class ChatGPTClient:
    def __init__(self):
        """Initialize the main window and OpenAI client."""
        self.window = ctk.CTk()
        self.window.title("ChatGPT Desktop")

        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        init_width = min(int(screen_width * 0.9), 1200)
        init_height = min(int(screen_height * 0.9), 800)
        self.window.geometry(f"{init_width}x{init_height}")
        # Allow shrinking on smaller displays
        self.window.minsize(800, 600)
        try:
            self.window.iconbitmap("@" + ICON_PATH)
        except Exception:
            logging.warning("Failed to set window icon")

        # モデルの初期値を環境変数から読み込む
        default_model = os.getenv("OPENAI_MODEL", "gpt-4.1-mini-2025-04-14")
        self.model_var = ctk.StringVar(value=default_model)
        
        # OpenAI クライアントの初期化
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            messagebox.showerror("エラー", "環境変数 OPENAI_API_KEY が設定されていません")
            logging.error("OPENAI_API_KEY is not set")
            self.window.destroy()
            return

        logging.info("Loaded OpenAI API key from environment")

        base_url = os.getenv("OPENAI_BASE_URL")
        if base_url:
            self.client = OpenAI(api_key=api_key, base_url=base_url)
        else:
            self.client = OpenAI(api_key=api_key)

        timeout_str = os.getenv("OPENAI_TIMEOUT", "0")
        try:
            self.timeout = float(timeout_str) or None
        except ValueError:
            logging.warning("Invalid OPENAI_TIMEOUT=%s, using default", timeout_str)
            self.timeout = None

        # 会話履歴
        self.messages = []
        self.current_title = None
        self.memory = ConversationMemory()
        self.uploaded_files = []
        self.response_queue = queue.Queue()
        self.assistant_start = None
        self.tot_start = None
        self._diagram_path: str | None = None
        self._failed_mermaid_code: str | None = None
        self.agent_var = ctk.StringVar(value="chatgpt")
        self.tot_level_var = ctk.StringVar(value="LOW")
        self.agent_tools = [
            get_web_scraper(),
            get_sqlite_tool(),
            get_graphviz_tool(),
            get_mermaid_tool(),
        ]
        self.tools = [
            {
                "type": "function",
                "function": {
                    "name": "create_graphviz_diagram",
                    "description": "DOT言語から図を生成する。フローチャート等に適している。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "code": {"type": "string", "description": "DOT言語のコード"}
                        },
                        "required": ["code"],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "create_mermaid_diagram",
                    "description": "Mermaid markdown-like codeから図を生成する。シーケンス図、ガントチャート等に適している。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "code": {"type": "string", "description": "Mermaid記法のコード"}
                        },
                        "required": ["code"],
                    },
                },
            },
        ]
        
        # UI要素の作成
        self.setup_ui()
        # キュー監視処理を開始
        self.window.after(100, self.process_queue)
        
    def setup_ui(self):
        """Build all widgets and configure layout."""
        # メインコンテナ
        main_container = ctk.CTkFrame(self.window, fg_color="transparent")
        main_container.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_columnconfigure(1, weight=1)

        # 右側のチャット/ヘルプ切り替えタブ
        tabview = ctk.CTkTabview(main_container)
        tabview.grid(row=0, column=1, sticky="nsew")
        chat_tab = tabview.add("チャット")
        info_tab = tabview.add("エージェント比較")
        
        # 左側パネル（設定）
        left_panel = ctk.CTkScrollableFrame(
            main_container,
            width=320,
            fg_color="#F1F3F4",
            corner_radius=8,
            border_width=0,
        )
        left_panel.grid(row=0, column=0, sticky="ns", padx=(0, 15))
        # CustomTkinter's CTkScrollableFrame uses grid layout; grid_propagate
        # takes no arguments, so call it without parameters to disable resizing
        left_panel.grid_propagate()
        
        # 設定タイトル
        settings_label = ctk.CTkLabel(left_panel, text="設定",
                                     font=(FONT_FAMILY, 22, "bold"))
        settings_label.pack(pady=20)
        
        # モデル選択
        model_label = ctk.CTkLabel(left_panel, text="モデル",
                                  font=(FONT_FAMILY, 16))
        model_label.pack(pady=(20, 5))
        
        model_menu = ctk.CTkOptionMenu(
            left_panel,
            values=[
                "gpt-3.5-turbo",
                "gpt-4o",
                "gpt-4.1-mini-2025-04-14",
                "gpt-4.1-nano-2025-04-14",
                "gpt-4.1-2025-04-14",
            ],
            variable=self.model_var,
            width=250,
        )
        model_menu.pack(pady=(0, 20))
        
        # 温度設定
        temp_label = ctk.CTkLabel(left_panel, text="温度: 0.7",
                                 font=(FONT_FAMILY, 16))
        temp_label.pack(pady=(20, 5))

        self.temp_slider = ctk.CTkSlider(left_panel, from_=0, to=2, number_of_steps=20,
                                        command=lambda v: temp_label.configure(text=f"温度: {v:.1f}"))
        self.temp_slider.set(0.7)
        self.temp_slider.pack(pady=(0, 20))

        agent_label = ctk.CTkLabel(left_panel, text="エージェント",
                                   font=(FONT_FAMILY, 16))
        agent_label.pack(pady=(10, 5))

        agent_menu = ctk.CTkOptionMenu(
            left_panel,
            values=["chatgpt", "react", "cot", "tot", "プレゼンテーション"],
            variable=self.agent_var,
            width=250,
        )
        agent_menu.pack(pady=(0, 20))

        tot_label = ctk.CTkLabel(left_panel, text="思考の木レベル",
                                   font=(FONT_FAMILY, 16))
        tot_label.pack(pady=(10, 5))

        tot_menu = ctk.CTkOptionMenu(
            left_panel,
            values=list(TOT_LEVELS.keys()),
            variable=self.tot_level_var,
            width=250,
        )
        tot_menu.pack(pady=(0, 20))
        
        # ファイルアップロードボタン
        upload_btn = ctk.CTkButton(left_panel, text="ファイルをアップロード",
                                  command=self.upload_file,
                                  font=(FONT_FAMILY, 16))
        upload_btn.pack(pady=10)
        
        # アップロードされたファイルリスト
        self.file_list_label = ctk.CTkLabel(left_panel, text="アップロードされたファイル:",
                                           font=(FONT_FAMILY, 14))
        self.file_list_label.pack(pady=(20, 5))
        
        self.file_list_text = ctk.CTkTextbox(left_panel, height=100, width=250)
        self.file_list_text.pack(pady=(0, 20))
        
        # 新しい会話ボタン
        new_chat_btn = ctk.CTkButton(left_panel, text="新しい会話",
                                    command=self.new_chat,
                                    font=(FONT_FAMILY, 16))
        new_chat_btn.pack(pady=10)

        load_chat_btn = ctk.CTkButton(
            left_panel,
            text="会話を読み込み",
            command=self.load_chat,
            font=(FONT_FAMILY, 16),
        )
        load_chat_btn.pack(pady=10)

        save_chat_btn = ctk.CTkButton(
            left_panel,
            text="会話を保存",
            command=self.save_conversation,
            font=(FONT_FAMILY, 16),
        )
        save_chat_btn.pack(pady=10)
        
        # 右側パネル（図プレビュー）
        self.diagram_panel = DiagramFrame(
            main_container,
            width=270,
            fg_color=None,
            corner_radius=8,
            border_width=0,
        )
        self.diagram_panel.grid(row=0, column=2, sticky="ns", padx=(15, 0))
        # Disable geometry propagation so the frame keeps the set width
        self.diagram_panel.grid_propagate()

        self.diagram_label = ctk.CTkLabel(self.diagram_panel, text="図のプレビュー", font=(FONT_FAMILY, 16))
        self.diagram_label.pack(padx=10, pady=10)

        self.save_button = ctk.CTkButton(
            self.diagram_panel,
            text="保存",
            command=lambda: self.save_diagram(),
            font=(FONT_FAMILY, 14),
            state="disabled",
        )
        self.save_button.pack(pady=(0, 10))

        self.copy_button = ctk.CTkButton(
            self.diagram_panel,
            text="コピー",
            command=lambda: self.copy_diagram(),
            font=(FONT_FAMILY, 14),
            state="disabled",
        )
        self.copy_button.pack(pady=(0, 10))

        self.clear_button = ctk.CTkButton(
            self.diagram_panel,
            text="クリア",
            command=lambda: self.clear_diagram(),
            font=(FONT_FAMILY, 14),
            state="disabled",
        )
        self.clear_button.pack(pady=(0, 10))

        self.fix_button = ctk.CTkButton(
            self.diagram_panel,
            text="修正",
            command=lambda: self.retry_diagram(),
            font=(FONT_FAMILY, 14),
            state="disabled",
        )
        self.fix_button.pack(pady=(0, 10))

        # 右側パネル（チャット）
        right_panel = ctk.CTkFrame(
            chat_tab,
            fg_color="#FFFFFF",
            corner_radius=8,
            border_width=0,
        )
        right_panel.grid(row=0, column=0, sticky="nsew")
        chat_tab.grid_rowconfigure(0, weight=1)
        chat_tab.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(0, weight=1)
        right_panel.grid_columnconfigure(0, weight=1)
        
        # チャットエリア
        self.chat_display = ctk.CTkTextbox(
            right_panel,
            font=(FONT_FAMILY, 16),
            wrap="word",
            fg_color="#FFFFFF",
        )
        self.chat_display.grid(row=0, column=0, sticky="nsew", padx=20, pady=(20, 10))
        self.chat_display.tag_config("user_msg", background="#FFFFFF")
        self.chat_display.tag_config("assistant_msg", background="#F1F3F4")
        
        # 入力エリア
        input_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        input_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 20))
        input_frame.grid_columnconfigure(0, weight=1)
        
        self.input_field = ctk.CTkEntry(
            input_frame,
            placeholder_text="メッセージを入力...",
            font=(FONT_FAMILY, 16),
            height=40,
        )
        self.input_field.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self.input_field.bind("<Return>", lambda e: self.send_message())
        
        send_btn = ctk.CTkButton(
            input_frame,
            text="送信",
            width=80,
            command=self.send_message,
            font=(FONT_FAMILY, 16),
        )
        send_btn.grid(row=0, column=1)

        # thinking indicator
        self.progress = ctk.CTkProgressBar(
            input_frame,
            mode="indeterminate",
        )
        self.progress.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(5, 0))
        self.progress.stop()

        help_box = ctk.CTkTextbox(info_tab, font=(FONT_FAMILY, 14), wrap="word")
        help_box.pack(fill="both", expand=True, padx=10, pady=10)
        help_text = (
            "# エージェント比較\n\n"
            "## Chain of Thought (CoT)\n"
            "1. 単純な線形推論を行います。\n"
            "2. '思考:' と '最終的な答え:' を順番に返します。\n"
            "例: 質問『富士山の高さは？』\n"
            "思考: 富士山は日本の最高峰で3776m。\n"
            "最終的な答え: 3776メートルです。\n\n"
            "## ReAct\n"
            "1. 思考とツール利用を交互に行います。\n"
            "2. '思考 -> 行動 -> 観察' を繰り返し最終回答を導きます。\n"
            "例: ウェブ検索を使って最新ニュースを取得。\n\n"
            "## Tree of Thoughts (ToT)\n"
            "1. 複数の思考パスを探索し最良の答えを選択します。\n"
            "2. 探索深さと分岐数は環境変数で調整できます。\n"
            "例: パズルの解法候補をいくつか試す。\n\n"
            "## プロンプト作成のコツ\n"
            "- 目的や条件を具体的に書く\n"
            "- 望ましい出力形式を明示する\n"
            "- 背景情報や制約を事前に伝える\n"
            "詳細は OpenAI のベストプラクティス(https://platform.openai.com/docs/guides/gpt-best-practices) を参照してください。"
        )
        help_box.insert("1.0", help_text)
        help_box.configure(state="disabled")
        
    def upload_file(self):
        """Prompt for a file and store its contents."""
        file_path = filedialog.askopenfilename(
            title="ファイルを選択",
            filetypes=[
                ("対応ファイル", "*.docx *.pdf *.png *.jpg *.jpeg *.xlsx"),
                ("Word", "*.docx"),
                ("PDF", "*.pdf"),
                ("画像", "*.png *.jpg *.jpeg"),
                ("Excel", "*.xlsx")
            ]
        )
        
        if file_path:
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            
            try:
                content = self.process_file(file_path, file_ext)
                self.uploaded_files.append({
                    "name": file_name,
                    "type": file_ext,
                    "content": content
                })
                
                # ファイルリストを更新
                self.update_file_list()
                
                messagebox.showinfo("成功", f"{file_name} をアップロードしました")
                
            except Exception as e:
                messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました: {str(e)}")
    
    def process_file(self, file_path: str, file_ext: str) -> str:
        """ファイルタイプに応じて内容を処理"""
        if file_ext == ".docx":
            doc = docx.Document(file_path)
            return "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
        elif file_ext == ".pdf":
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    extracted_text = page.extract_text()
                    if extracted_text:
                        text += extracted_text + "\n"
                return text
                
        elif file_ext in [".png", ".jpg", ".jpeg"]:
            # 画像をbase64エンコード
            with open(file_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode('utf-8')
                
        elif file_ext == ".xlsx":
            workbook = openpyxl.load_workbook(file_path)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append(list(row))
                
                sheets_data[sheet_name] = sheet_data
            
            # シート情報を文字列として返す
            result = f"Excelファイル: {len(workbook.sheetnames)}個のシート\n"
            for sheet_name, data in sheets_data.items():
                result += f"\n【シート: {sheet_name}】\n"
                result += f"行数: {len(data)}\n"
                if data:
                    result += f"列数: {len(data[0])}\n"
                    # 最初の数行を表示
                    for i, row in enumerate(data[:5]):
                        result += f"行{i+1}: {row}\n"
                    if len(data) > 5:
                        result += "...\n"

            workbook.close()
            return result
        
        return ""
    
    def update_file_list(self):
        """アップロードされたファイルリストを更新"""
        self.file_list_text.configure(state="normal")
        self.file_list_text.delete("1.0", "end")
        for file in self.uploaded_files:
            self.file_list_text.insert("end", f"• {file['name']}\n")
        self.file_list_text.configure(state="disabled")

    def adjust_width_for_message(self, msg: str) -> None:
        """Expand the window if a long message is entered."""
        if not hasattr(self, "window") or not hasattr(self, "chat_display"):
            return
        try:
            char_len = len(msg)
            font_desc = self.chat_display.cget("font")
            import tkinter.font as tkfont
            fnt = tkfont.Font(font=font_desc)
            char_len = fnt.measure(msg)
        except Exception:
            pass

        threshold = 1000
        cur_width = self.window.winfo_width()
        screen_width = self.window.winfo_screenwidth()
        if char_len > threshold and cur_width < screen_width:
            step = 150
            new_width = min(cur_width + step, screen_width)
            height = self.window.winfo_height()
            self.window.geometry(f"{new_width}x{height}")
    
    def send_message(self):
        """Handle user input and start fetching a reply."""
        user_message = self.input_field.get().strip()
        if not user_message:
            return

        # メッセージをクリア
        self.input_field.delete(0, "end")

        if hasattr(self, "progress"):
            try:
                self.progress.stop()
                self.progress.start()
            except Exception:
                pass
        
        # ユーザーメッセージを表示
        self.chat_display.configure(state="normal")
        start = self.chat_display.index("end") if hasattr(self.chat_display, "index") else None
        self.chat_display.insert("end", f"\nYou: {user_message}\n\n")
        if start is not None and hasattr(self.chat_display, "tag_add"):
            end = self.chat_display.index("end")
            if hasattr(self.chat_display, "tag_remove"):
                self.chat_display.tag_remove("assistant_msg", start, end)
            self.chat_display.tag_add("user_msg", start, end)
        self.chat_display.see("end")
        self.chat_display.configure(state="disabled")
        self.adjust_width_for_message(user_message)
        
        # ファイル情報を含めたメッセージを作成
        # OpenAI APIは、"user"ロールのメッセージcontentに直接画像を含めることを想定
        # "vision"モデル (例: gpt-4-vision-preview, gpt-4o) は content に配列を受け付けます。
        # ここでは、ファイルの内容をテキストとして含めることを前提としています。
        # 画像ファイルがある場合、その内容 (base64) は直接メッセージに含めず、
        # 別途、GPT-4V (Vision) などのマルチモーダルモデルのAPIコール時に適切に処理する必要があります。
        # 現在のコードでは、画像はbase64エンコードされた文字列としてcontentに含めていますが、
        # これが直接的にテキストモデルで解釈されるわけではありません。
        # gpt-4oのようなマルチモーダルモデルでは、contentに画像データを含めるための特定の形式が必要です。
        # (例: `{"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_string}"}}`)
        # 今回の修正はモデル名の変更のみに留め、このロジックは変更しません。

        content_parts = [{"type": "text", "text": user_message}]

        if not self.messages:
            system_prompt = (
                "あなたは優秀なAIアシスタントです。"
                "ユーザーの質問が曖昧だと判断した場合、"
                "OpenAI が公開しているプロンプト作成ベストプラクティス(https://platform.openai.com/docs/guides/gpt-best-practices)を参考に、"
                "通常の回答の後で『【プロンプトアドバイス】』という見出しを付け、"
                "より具体的に質問するための例を2〜3個日本語で提示してください。"
                "例: 『ラーメンについて教えて』→『東京でおすすめの醤油ラーメンのお店は？』"
            )
            self.messages.append({"role": "system", "content": system_prompt})
        
        if self.uploaded_files:
            file_info_text = "\n\n【アップロードされたファイル情報】\n"
            for file in self.uploaded_files:
                if file['type'] in ['.docx', '.pdf', '.xlsx']:
                    file_info_text += f"\n--- {file['name']} ---\n{file['content'][:1000]}...\n" # 長すぎる内容は省略
                elif file['type'] in ['.png', '.jpg', '.jpeg']:
                    # gpt-4oなどのマルチモーダルモデルの場合、画像は特別な形式で渡す
                    # ここでは、メッセージに追加のテキスト情報としてファイル名のみ含めるか、
                    # あるいは、content_partsに画像データを追加する処理が必要。
                    # 今回のモデル名変更のリクエストでは、この部分のロジックは変更しない。
                    # 単純にテキストとしてファイル名を付加する例：
                    file_info_text += f"\n画像ファイル: {file['name']} (内容は別途送信されます)\n"
                    # もしgpt-4oに画像を直接渡すなら、以下のような形式でcontent_partsに追加
                    # image_data = {
                    #     "type": "image_url",
                    #     "image_url": {
                    #         "url": f"data:image/{file['type'][1:]};base64,{file['content']}"
                    #     }
                    # }
                    # content_parts.append(image_data)

            # ユーザーメッセージのテキストパートにファイル情報を追加
            content_parts[0]["text"] += file_info_text
        
        # メッセージを履歴に追加
        self.messages.append({"role": "user", "content": content_parts})
        
        # 初回メッセージの場合、タイトルを生成
        user_count = sum(1 for m in self.messages if m.get("role") == "user")
        if user_count == 1:
            self.generate_title(user_message)
        
        # エージェント種別に応じて応答を取得
        if getattr(self, "agent_var", None) and self.agent_var.get() != "chatgpt":
            agent_type = self.agent_var.get()
            threading.Thread(target=self.run_agent, args=(agent_type, user_message), daemon=True).start()
        else:
            threading.Thread(target=self.get_response, daemon=True).start()
    
    def get_response(self):
        """Stream the assistant's reply, execute tool calls, and push updates."""
        try:
            self.response_queue.put("Assistant: ")
            while True:
                response_text = ""
                tool_data: dict[str, dict[str, str]] = {}

                params = {
                    "model": self.model_var.get(),
                    "messages": self.messages,
                    "temperature": self.temp_slider.get(),
                    "stream": True,
                }
                if getattr(self, "tools", None):
                    params["tools"] = self.tools
                    params["tool_choice"] = "auto"
                timeout_val = getattr(self, "timeout", None)
                if timeout_val is not None:
                    params["timeout"] = timeout_val

                stream = self.client.chat.completions.create(**params)
                finish_reason = None

                for chunk in stream:
                    choice = chunk.choices[0]
                    delta = choice.delta
                    finish_reason = choice.finish_reason

                    if getattr(delta, "content", None) is not None:
                        content = delta.content
                        response_text += content
                        self.response_queue.put(content)

                    calls = getattr(delta, "tool_calls", None)
                    if calls:
                        for c in calls:
                            info = tool_data.setdefault(c.id, {"name": "", "args": ""})
                            if getattr(c.function, "name", None):
                                info["name"] = c.function.name
                            if getattr(c.function, "arguments", None):
                                info["args"] += c.function.arguments

                self.response_queue.put("\n")

                if tool_data and finish_reason == "tool_calls":
                    assistant_msg = {
                        "role": "assistant",
                        "content": response_text,
                        "tool_calls": [
                            {
                                "id": cid,
                                "type": "function",
                                "function": {"name": d["name"], "arguments": d["args"]},
                            }
                            for cid, d in tool_data.items()
                        ],
                    }
                    self.messages.append(assistant_msg)

                    for cid, d in tool_data.items():
                        func = TOOL_FUNCS.get(d["name"])
                        if func:
                            try:
                                args = json.loads(d["args"] or "{}")
                                if d["name"] == "create_mermaid_diagram":
                                    code = args.get("code", "")
                                    self._failed_mermaid_code = sanitize_mermaid_code(code)
                                result = func(**args)
                            except Exception as exc:
                                result = f"Tool execution failed: {exc}"
                        else:
                            result = f"Unknown tool: {d['name']}"
                        self.messages.append({"role": "tool", "tool_call_id": cid, "content": result})

                    # Continue looping to stream assistant's follow-up answer
                    continue

                # 通常の応答を保存して終了
                self.messages.append({"role": "assistant", "content": response_text})
                match = re.search(r"(?:[A-Za-z]:)?[\\/][^\s]+\.png", response_text)
                if match and os.path.isfile(match.group(0)):
                    self.response_queue.put(f"__DIAGRAM__{match.group(0)}")
                self.response_queue.put("__SAVE__")
                break

        except Exception as e:
            logging.exception("Streaming failed: %s", e)
            self.response_queue.put(f"\n\nエラー: {str(e)}\n")

    def simple_llm(self, prompt: str, *, stream: bool = False, prefix: str = "") -> str:
        """Call the OpenAI API and optionally stream tokens to the queue."""
        params = {
            "model": self.model_var.get(),
            "messages": [{"role": "user", "content": prompt}],
            "temperature": self.temp_slider.get(),
        }
        timeout_val = getattr(self, "timeout", None)
        if timeout_val is not None:
            params["timeout"] = timeout_val
        if stream:
            params["stream"] = True
            result = ""
            try:
                for chunk in self.client.chat.completions.create(**params):
                    delta = chunk.choices[0].delta
                    if getattr(delta, "content", None) is not None:
                        text = delta.content
                        result += text
                        self.response_queue.put(prefix + text)
            except Exception as exc:
                logging.exception("Streaming call failed: %s", exc)
            return result
        resp = self.client.chat.completions.create(**params)
        return resp.choices[0].message.content

    def run_agent(self, agent_type: str, question: str) -> None:
        """Execute the selected agent and stream steps to the queue."""
        try:
            self.response_queue.put("Assistant: ")
            if agent_type == "react":
                agent = ReActAgent(
                    functools.partial(self.simple_llm, stream=True),
                    self.agent_tools,
                    self.memory,
                )
            elif agent_type == "cot":
                agent = CoTAgent(functools.partial(self.simple_llm, stream=True), self.memory)
            elif agent_type == "tot":
                level = self.tot_level_var.get()
                depth, breadth = TOT_LEVELS.get(level, (2, 2))
                try:
                    env_depth, env_breadth = read_tot_env()
                    if env_depth is not None:
                        depth = env_depth
                    if env_breadth is not None:
                        breadth = env_breadth
                except (SystemExit, Exception) as exc:
                    self.response_queue.put(f"\n\nエラー: {exc}\n")
                    return
                evaluator = create_evaluator(self.simple_llm)
                agent = ToTAgent(
                    functools.partial(self.simple_llm, stream=True, prefix="__TOT__"),
                    evaluator,
                    max_depth=depth,
                    breadth=breadth,
                    memory=self.memory,
                )
            elif agent_type == "プレゼンテーション":
                agent = PresentationAgent(self.simple_llm)
            else:
                self.response_queue.put("未対応のエージェントです\n")
                return
            if agent_type == "tot":
                self.response_queue.put("__TOT_START__")
                final_answer = ""
                for step in agent.run_iter(question):
                    if step.startswith("最終的な答え:"):
                        final_answer = step[len("最終的な答え:"):].strip()
                self.response_queue.put("__TOT_END__" + final_answer + "\n")
                self.messages.append({"role": "user", "content": question})
                self.messages.append({"role": "assistant", "content": final_answer})
                match = re.search(r"(?:[A-Za-z]:)?[\\/][^\"\n]+\.png", final_answer)
                if match and os.path.isfile(match.group(0)):
                    self.response_queue.put(f"__DIAGRAM__{match.group(0)}")
                self.response_queue.put("__SAVE__")
            else:
                response_text = ""
                for step in agent.run_iter(question):
                    response_text += step + "\n"
                self.messages.append({"role": "user", "content": question})
                self.messages.append({"role": "assistant", "content": response_text})
                match = re.search(r"(?:[A-Za-z]:)?[\\/][^\"\n]+\.png", response_text)
                if match and os.path.isfile(match.group(0)):
                    self.response_queue.put(f"__DIAGRAM__{match.group(0)}")
                self.response_queue.put("__SAVE__")
        except Exception as exc:
            self.response_queue.put(f"\n\nエラー: {exc}\n")
    
    def generate_title(self, first_message: str):
        """最初のメッセージからタイトルを生成"""
        try:
            # タイトル生成はシンプルなモデルで十分
            params = {
                "model": "gpt-3.5-turbo",  # タイトル生成用モデルは変更なし
                "messages": [
                    {
                        "role": "system",
                        "content": "ユーザーのメッセージから、短く簡潔な会話のタイトルを生成してください。20文字以内で。",
                    },
                    {"role": "user", "content": first_message},
                ],
                "temperature": 0.5,
                "max_tokens": 50,
            }
            timeout_val = getattr(self, "timeout", None)
            if timeout_val is not None:
                params["timeout"] = timeout_val
            response = self.client.chat.completions.create(**params)

            self.current_title = response.choices[0].message.content.strip()
            self.response_queue.put(f"__TITLE__{self.current_title}")
            
        except Exception:
            logging.exception("Failed to generate title")
            self.current_title = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.response_queue.put(f"__TITLE__{self.current_title}")
    
    def save_conversation(self, show_popup: bool = True):
        """会話をJSONファイルとして保存."""
        if not self.current_title:
            return
        
        # conversationsディレクトリがなければ作成
        if not os.path.exists(CONV_DIR):
            os.makedirs(CONV_DIR)
            
        filename_base = f"{self.current_title}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        # ファイル名に使えない文字を置換
        filename_safe = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filename_base)
        filename = os.path.join(CONV_DIR, f"{filename_safe}.json")
        
        # uploaded_filesのcontentは保存しない (大きすぎる可能性があるため)
        files_metadata = []
        for f_info in self.uploaded_files:
            files_metadata.append({
                "name": f_info["name"],
                "type": f_info["type"]
            })

        conversation_data = {
            "title": self.current_title,
            "timestamp": datetime.datetime.now().isoformat(),
            "model": self.model_var.get(),
            "messages": self.messages, # メッセージ履歴をそのまま保存
            "uploaded_files_metadata": files_metadata # contentは含めない
        }
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(conversation_data, f, ensure_ascii=False, indent=2)
            if show_popup:
                try:
                    messagebox.showinfo("保存完了", f"会話を {filename} に保存しました")
                except tkinter.TclError:
                    pass
        except Exception as e:
            if show_popup:
                messagebox.showerror("保存エラー", f"会話の保存に失敗しました: {str(e)}")
            else:
                logging.error("会話の保存に失敗しました: %s", e)

    
    def new_chat(self):
        """新しい会話を開始"""
        self.messages = []
        self.current_title = None
        self.uploaded_files = []
        try:
            if hasattr(self, "memory") and hasattr(self.memory, "clear"):
                self.memory.clear()
        except Exception:
            logging.warning("Failed to reset memory", exc_info=True)

        self.chat_display.configure(state="normal")
        self.chat_display.delete("1.0", "end")
        self.chat_display.insert("1.0", "新しい会話を開始しました。\n")
        self.chat_display.configure(state="disabled")
        
        self.file_list_text.configure(state="normal")
        self.file_list_text.delete("1.0", "end")
        self.file_list_text.configure(state="disabled")

        # 既存の図プレビューをリセット
        try:
            self.clear_diagram()
        except Exception:
            logging.warning("Failed to clear diagram", exc_info=True)

        self.window.title("ChatGPT Desktop")

    def load_chat(self):
        """Open a saved conversation file and load its content."""
        file_path = filedialog.askopenfilename(
            title="会話を選択",
            filetypes=[("Conversation", "*.json")],
            initialdir=CONV_DIR,
        )
        if file_path:
            self.load_conversation(file_path)

    def load_conversation(self, file_path: str):
        """Load conversation from a JSON file created by save_conversation."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("読み込みエラー", f"会話の読み込みに失敗しました: {str(e)}")
            return

        self.current_title = data.get("title")
        if self.current_title:
            self.window.title(f"ChatGPT Desktop - {self.current_title}")
        self.messages = data.get("messages", [])
        meta = data.get("uploaded_files_metadata", [])
        self.uploaded_files = [{"name": m["name"], "type": m["type"]} for m in meta]

        # Refresh displays
        self.update_file_list()
        self.chat_display.configure(state="normal")
        self.chat_display.delete("1.0", "end")
        for msg in self.messages:
            role = msg.get("role")
            content = msg.get("content", "")
            if isinstance(content, list):
                # content may be structured as list of parts
                content = "".join(part.get("text", "") for part in content)
            prefix = "You" if role == "user" else "Assistant"
            start = self.chat_display.index("end") if hasattr(self.chat_display, "index") else None
            self.chat_display.insert("end", f"\n{prefix}: {content}\n\n")
            if start is not None and hasattr(self.chat_display, "tag_add"):
                end = self.chat_display.index("end")
                tag = "user_msg" if role == "user" else "assistant_msg"
                if role == "user" and hasattr(self.chat_display, "tag_remove"):
                    self.chat_display.tag_remove("assistant_msg", start, end)
                self.chat_display.tag_add(tag, start, end)
        self.chat_display.configure(state="disabled")

    def display_diagram(self, path: str) -> None:
        """Preview a diagram PNG and enable saving."""
        try:
            img = Image.open(path)
            preview = ctk.CTkImage(light_image=img, size=(200, 200))
        except Exception:
            logging.exception("Failed to load diagram %s", path)
            return
        self.diagram_label.configure(image=preview, text="")
        self.diagram_label.image = preview
        self.save_button.configure(state="normal")
        self.clear_button.configure(state="normal")
        self.copy_button.configure(state="normal")
        self.fix_button.configure(state="disabled")
        self._diagram_path = path

    def save_diagram(self) -> None:
        """Save the currently previewed diagram to a location chosen by the user."""
        if not getattr(self, "_diagram_path", None):
            return
        dest = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG", "*.png")])
        if dest:
            try:
                shutil.copy(self._diagram_path, dest)
            except Exception as exc:
                messagebox.showerror("保存エラー", str(exc))
            else:
                messagebox.showinfo("保存完了", f"図を {dest} に保存しました")

    def clear_diagram(self) -> None:
        """Remove the current diagram preview and disable related buttons."""
        self.diagram_label.configure(image=None, text="図のプレビュー")
        self.diagram_label.image = None
        self.save_button.configure(state="disabled")
        self.clear_button.configure(state="disabled")
        self.copy_button.configure(state="disabled")
        self.fix_button.configure(state="disabled")
        self._diagram_path = None

    def copy_diagram(self) -> None:
        """Copy the diagram path to the clipboard."""
        if not getattr(self, "_diagram_path", None):
            return
        try:
            self.window.clipboard_clear()
            self.window.clipboard_append(self._diagram_path)
            messagebox.showinfo("コピー完了", "図のファイルパスをコピーしました")
        except Exception as exc:
            messagebox.showerror("コピーエラー", str(exc))

    def retry_diagram(self) -> None:
        """Attempt to regenerate a Mermaid diagram after a failure."""
        code = getattr(self, "_failed_mermaid_code", None)
        if not code:
            return
        path = create_mermaid_diagram(code)
        if path.startswith("Failed to generate diagram"):
            messagebox.showerror("エラー", "図の生成に失敗しました")
            return
        self.display_diagram(path)
        self.fix_button.configure(state="disabled")

    def process_queue(self):
        """キューからのメッセージをGUIに反映"""
        try:
            while True:
                item = self.response_queue.get_nowait()
                if item.startswith("__TITLE__"):
                    title = item[len("__TITLE__"):]
                    self.window.title(f"ChatGPT Desktop - {title}")
                    continue
                if item.startswith("__DIAGRAM__"):
                    self.display_diagram(item[len("__DIAGRAM__"):])
                    continue
                if item.startswith("Failed to generate diagram:"):
                    self.diagram_label.configure(image=None, text="図の生成に失敗しました")
                    self.fix_button.configure(state="normal")
                    self.clear_button.configure(state="normal")
                    self.save_button.configure(state="disabled")
                    self.copy_button.configure(state="disabled")
                    self._diagram_path = None
                    # self._failed_mermaid_code is already set during tool call
                    continue
                if item == "__SAVE__":
                    threading.Thread(
                        target=self.save_conversation,
                        kwargs={"show_popup": False},
                        daemon=True,
                    ).start()
                    if hasattr(self, "progress"):
                        try:
                            self.progress.stop()
                        except Exception:
                            pass
                    continue
                if item == "__TOT_START__":
                    self.tot_start = None
                    continue
                self.chat_display.configure(state="normal")
                if item.startswith("Assistant: "):
                    self.assistant_start = self.chat_display.index("end") if hasattr(self.chat_display, "index") else None
                    self.chat_display.insert("end", item)
                elif item.startswith("__TOT__"):
                    if self.tot_start is None:
                        self.tot_start = self.chat_display.index("end")
                    self.chat_display.insert("end", item[len("__TOT__"):])
                elif item.startswith("__TOT_END__"):
                    final = item[len("__TOT_END__"):]
                    if self.tot_start is not None:
                        self.chat_display.delete(self.tot_start, "end")
                        self.tot_start = None
                    self.chat_display.insert("end", final)
                    if (
                        self.assistant_start is not None
                        and final.endswith("\n")
                        and hasattr(self.chat_display, "tag_add")
                    ):
                        end = self.chat_display.index("end")
                        self.chat_display.tag_add(
                            "assistant_msg", self.assistant_start, end
                        )
                        self.assistant_start = None
                else:
                    self.chat_display.insert("end", item)
                    if self.assistant_start is not None and item.endswith("\n") and hasattr(self.chat_display, "tag_add"):
                        end = self.chat_display.index("end")
                        self.chat_display.tag_add("assistant_msg", self.assistant_start, end)
                        self.assistant_start = None
                if "エラー" in item and hasattr(self, "progress"):
                    try:
                        self.progress.stop()
                    except Exception:
                        pass
                self.chat_display.see("end")
                self.chat_display.configure(state="disabled")
        except queue.Empty:
            pass
        self.window.after(100, self.process_queue)
    
    def run(self):
        """Start the application event loop."""
        # 初期化時にチャット表示とファイルリストをdisabledに
        self.chat_display.configure(state="disabled")
        self.file_list_text.configure(state="disabled")
        self.window.mainloop()

if __name__ == "__main__":
    app = ChatGPTClient()
    app.run()
